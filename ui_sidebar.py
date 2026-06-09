import json
import os
import tempfile
from datetime import datetime

import customtkinter as ctk

from ui_constants import (
    SIDEBAR_BG, BG_MAIN, AI_BG, ACCENT, ACCENT_DIM, ACCENT_HOV,
    TEXT_PRI, TEXT_SEC, TEXT_OK,
    BORDER, DIVIDER, BTN_IDLE, BTN_HOV, BTN_BORDER,
    SIDEBAR_W, FONT_BADGE, FONT_SMALL, FONT_MICRO,
    _font, _F_UI,
)
from ui_widgets import _HoverButton
from core_engine import trova_modelli, _APP_ROOT


class SidebarMixin:
    """Costruzione sidebar, cronologia persistente ed export PDF."""

    # ── Costruzione sidebar ───────────────────────────────────────────────────

    def _build_sidebar(self):
        sb = ctk.CTkFrame(self, width=SIDEBAR_W, fg_color=SIDEBAR_BG, corner_radius=0)
        sb.grid(row=0, column=0, sticky="nsew")
        sb.grid_propagate(False)
        sb.grid_columnconfigure(0, weight=1)
        sb.grid_rowconfigure(19, weight=1)

        # Logo
        logo_frame = ctk.CTkFrame(sb, fg_color="transparent")
        logo_frame.grid(row=0, column=0, padx=20, pady=(28, 6), sticky="ew")
        ctk.CTkLabel(logo_frame, text="◈",
                     font=(_F_UI, 20, "bold"), text_color=ACCENT).pack(side="left", padx=(0, 8))
        title_stack = ctk.CTkFrame(logo_frame, fg_color="transparent")
        title_stack.pack(side="left")
        ctk.CTkLabel(title_stack, text="Data-Whisperer",
                     font=_font(13, "bold"), text_color=TEXT_PRI, anchor="w").pack(anchor="w")
        ctk.CTkLabel(title_stack, text="Offline AI Analyst",
                     font=FONT_MICRO, text_color=TEXT_SEC, anchor="w").pack(anchor="w")

        self._divider(sb, row=1)

        # Stato modello
        self._lbl_stato = ctk.CTkLabel(
            sb, text="● Rilevamento hardware...",
            font=_font(11, "bold"), text_color="#FFA726",
            anchor="w", wraplength=SIDEBAR_W - 28)
        self._lbl_stato.grid(row=2, column=0, padx=20, pady=(16, 4), sticky="ew")
        self._lbl_hw = ctk.CTkLabel(
            sb, text="", font=FONT_SMALL, text_color=TEXT_SEC,
            anchor="w", wraplength=SIDEBAR_W - 28, justify="left")
        self._lbl_hw.grid(row=3, column=0, padx=20, pady=(0, 12), sticky="ew")

        self._divider(sb, row=4)

        # Selettore Modello
        ctk.CTkLabel(sb, text="MODELLO", font=FONT_BADGE,
                     text_color=TEXT_SEC, anchor="w").grid(
            row=5, column=0, padx=20, pady=(14, 6), sticky="ew")

        _modelli      = trova_modelli()
        _nomi         = [os.path.basename(p) for p in _modelli] or ["Nessun modello installato"]
        _nome_default = os.path.basename(self._modello_path) if self._modello_path else _nomi[0]
        if _nome_default not in _nomi:
            _nome_default = _nomi[0]

        self._var_modello = ctk.StringVar(value=_nome_default)
        self._opt_modello = ctk.CTkOptionMenu(
            sb, values=_nomi, variable=self._var_modello,
            command=self._on_modello_selezionato,
            fg_color=BTN_IDLE, button_color=ACCENT_DIM, button_hover_color=ACCENT,
            text_color=TEXT_PRI, dropdown_fg_color=SIDEBAR_BG,
            dropdown_text_color=TEXT_PRI, dropdown_hover_color=BTN_HOV,
            font=_font(10), height=34, corner_radius=10,
        )
        self._opt_modello.grid(row=6, column=0, padx=16, pady=(0, 8), sticky="ew")

        _HoverButton(sb, text="   🗄️  Hub Modelli", height=34,
                     font=_font(11, "bold"), fg_color=BTN_IDLE, hover_color=BTN_HOV,
                     text_color=TEXT_SEC, border_color=BTN_BORDER, border_width=1,
                     corner_radius=10, command=self._apri_hub, anchor="w",
                     ).grid(row=7, column=0, padx=16, pady=(0, 14), sticky="ew")

        self._divider(sb, row=8)

        # Dataset
        ctk.CTkLabel(sb, text="DATASET", font=FONT_BADGE,
                     text_color=TEXT_SEC, anchor="w").grid(
            row=9, column=0, padx=20, pady=(14, 6), sticky="ew")

        self._btn_file = _HoverButton(
            sb, text="   ＋  Carica CSV / Excel", height=40,
            font=_font(12, "bold"), fg_color=BTN_IDLE, hover_color=BTN_HOV,
            text_color=ACCENT, border_color=BTN_BORDER, border_width=1,
            corner_radius=10, command=self._seleziona_file, state="disabled", anchor="w",
        )
        self._btn_file.grid(row=10, column=0, padx=16, pady=(0, 8), sticky="ew")

        self._lbl_file = ctk.CTkLabel(
            sb, text="Nessun file caricato", font=FONT_SMALL, text_color=TEXT_SEC,
            anchor="w", wraplength=SIDEBAR_W - 28, justify="left")
        self._lbl_file.grid(row=11, column=0, padx=20, pady=(0, 12), sticky="new")

        self._divider(sb, row=12)

        # Sessione
        ctk.CTkLabel(sb, text="SESSIONE", font=FONT_BADGE,
                     text_color=TEXT_SEC, anchor="w").grid(
            row=13, column=0, padx=20, pady=(14, 6), sticky="ew")

        _HoverButton(sb, text="   ↺  Nuova Chat", height=38,
                     font=_font(12, "bold"), fg_color=BTN_IDLE, hover_color=BTN_HOV,
                     text_color=TEXT_PRI, border_color=BTN_BORDER, border_width=1,
                     corner_radius=10, command=self._nuova_chat, anchor="w",
                     ).grid(row=14, column=0, padx=16, pady=(0, 6), sticky="ew")

        _HoverButton(sb, text="   📄  Esporta PDF", height=38,
                     font=_font(12, "bold"), fg_color=BTN_IDLE, hover_color=BTN_HOV,
                     text_color=TEXT_PRI, border_color=BTN_BORDER, border_width=1,
                     corner_radius=10, command=self._esporta_pdf_sessione, anchor="w",
                     ).grid(row=15, column=0, padx=16, pady=(0, 4), sticky="ew")

        _HoverButton(sb, text="   📊  Esporta XLSX", height=38,
                     font=_font(12, "bold"), fg_color=BTN_IDLE, hover_color=BTN_HOV,
                     text_color=TEXT_PRI, border_color=BTN_BORDER, border_width=1,
                     corner_radius=10, command=self._esporta_xlsx_sessione, anchor="w",
                     ).grid(row=16, column=0, padx=16, pady=(0, 12), sticky="ew")

        self._divider(sb, row=17)

        # Cronologia
        ctk.CTkLabel(sb, text="CRONOLOGIA", font=FONT_BADGE,
                     text_color=TEXT_SEC, anchor="w").grid(
            row=18, column=0, padx=20, pady=(12, 4), sticky="ew")

        self._storia_frame = ctk.CTkScrollableFrame(
            sb, fg_color="transparent",
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=ACCENT_DIM,
            corner_radius=0,
        )
        self._storia_frame.grid(row=19, column=0, sticky="nsew", padx=6)
        self._storia_frame.grid_columnconfigure(0, weight=1)

        self._divider(sb, row=20)

        # Footer
        footer = ctk.CTkFrame(sb, fg_color="transparent")
        footer.grid(row=21, column=0, padx=14, pady=(8, 16), sticky="ew")
        ctk.CTkLabel(footer, text="v1.0", font=FONT_MICRO, text_color=TEXT_SEC).pack(side="left")
        ctk.CTkLabel(footer, text="  Air-Gapped  ·  Zero rete",
                     font=FONT_MICRO, text_color=TEXT_SEC).pack(side="left")
        ctk.CTkButton(
            footer, text="?", width=24, height=24,
            font=_font(11, "bold"), fg_color=BTN_IDLE, hover_color=BTN_HOV,
            text_color=TEXT_SEC, border_color=BTN_BORDER, border_width=1,
            corner_radius=12, command=self._apri_tour,
        ).pack(side="right")

    def _divider(self, parent, row: int):
        ctk.CTkFrame(parent, height=1, fg_color=DIVIDER).grid(
            row=row, column=0, sticky="ew", padx=14)

    # ── Cronologia ────────────────────────────────────────────────────────────

    def _carica_storia(self):
        try:
            if os.path.exists(self._storia_path):
                with open(self._storia_path, "r", encoding="utf-8") as f:
                    self._history = json.load(f)
        except Exception:
            self._history = []

    def _salva_storia(self, query: str, output_testo: str = "",
                     grafici_nomi: "list | None" = None):
        entry = {
            "v":           2,
            "ts":          datetime.now().strftime("%d/%m %H:%M"),
            "dataset":     self._dataset.nome_file if self._dataset else "",
            "query":       query,
            "output_testo": output_testo,
            "grafici_nomi": grafici_nomi or [],
        }
        self._history.insert(0, entry)
        self._history = self._history[:50]
        try:
            with open(self._storia_path, "w", encoding="utf-8") as f:
                json.dump(self._history, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _aggiorna_storia_sidebar(self):
        if not hasattr(self, "_storia_frame"):
            return
        for w in self._storia_frame.winfo_children():
            w.destroy()
        for entry in self._history[:20]:
            has_chart  = bool(entry.get("grafici_nomi"))
            has_text   = bool((entry.get("output_testo") or "").strip())
            badge      = "📊 " if has_chart else ("📝 " if has_text else "")
            label      = entry["query"][:22] + ("…" if len(entry["query"]) > 22 else "")
            _HoverButton(
                self._storia_frame, text=f"{badge}{label}", height=26,
                font=_font(10), fg_color="transparent", hover_color=BTN_HOV,
                text_color=TEXT_SEC, border_width=0, corner_radius=6,
                command=lambda e=entry: self._mostra_dettaglio_query(e), anchor="w",
            ).pack(fill="x", padx=2, pady=1)

    def _mostra_dettaglio_query(self, entry: dict):
        from ui_constants import AI_BG, CODE_BG, BTN_IDLE, BTN_HOV, BTN_BORDER, TEXT_PRI
        dlg = ctk.CTkToplevel(self)
        dlg.title("Dettaglio analisi")
        dlg.configure(fg_color=AI_BG)
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.transient(self)

        w, h = 560, 520
        px = self.winfo_x() + (self.winfo_width()  - w) // 2
        py = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f"{w}x{h}+{px}+{py}")
        dlg.minsize(400, 300)

        scroll = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=0, pady=0)
        scroll.grid_columnconfigure(0, weight=1)

        def _sec_label(text):
            ctk.CTkLabel(scroll, text=text, font=_font(9, "bold"),
                         text_color=TEXT_SEC, anchor="w").pack(
                fill="x", padx=20, pady=(14, 2))

        # ── Metadata ────────────────────────────────────────────────────────
        meta = f"{entry.get('ts', '')}  ·  {entry.get('dataset', '')}"
        ctk.CTkLabel(scroll, text=meta, font=FONT_MICRO,
                     text_color=TEXT_SEC, anchor="w").pack(
            fill="x", padx=20, pady=(16, 2))

        # ── Query ───────────────────────────────────────────────────────────
        _sec_label("DOMANDA")
        q_box = ctk.CTkTextbox(scroll, height=60, font=_font(11),
                               fg_color=CODE_BG, text_color=TEXT_PRI,
                               wrap="word", activate_scrollbars=False)
        q_box.pack(fill="x", padx=16, pady=(0, 4))
        q_box.insert("1.0", entry.get("query", ""))
        q_box.configure(state="disabled")

        # ── Output testo ────────────────────────────────────────────────────
        output = (entry.get("output_testo") or "").strip()
        if output:
            _sec_label("RISPOSTA")
            out_box = ctk.CTkTextbox(scroll, height=100, font=_font(11),
                                     fg_color=CODE_BG, text_color=TEXT_PRI,
                                     wrap="word", activate_scrollbars=True)
            out_box.pack(fill="x", padx=16, pady=(0, 4))
            out_box.insert("1.0", output)
            out_box.configure(state="disabled")

        # ── Grafici thumbnail ────────────────────────────────────────────────
        grafici_nomi = entry.get("grafici_nomi") or []
        for nome in grafici_nomi:
            path = os.path.join(_APP_ROOT, "output", nome)
            if not os.path.exists(path):
                continue
            try:
                from PIL import Image as _PIL_Image
                img_pil = _PIL_Image.open(path)
                thumb_w = min(w - 60, 480)
                ratio   = thumb_w / img_pil.width
                thumb_h = int(img_pil.height * ratio)
                img_pil = img_pil.resize((thumb_w, thumb_h), _PIL_Image.LANCZOS)
                ctk_img = ctk.CTkImage(light_image=img_pil, dark_image=img_pil,
                                       size=(thumb_w, thumb_h))
                if not hasattr(self, "_dettaglio_img_refs"):
                    self._dettaglio_img_refs = []
                self._dettaglio_img_refs.append(ctk_img)
                ctk.CTkLabel(scroll, image=ctk_img, text="").pack(
                    padx=16, pady=(8, 4))
            except Exception:
                pass

        # ── Bottoni ────────────────────────────────────────────────────────
        bar = ctk.CTkFrame(dlg, fg_color="transparent")
        bar.pack(fill="x", padx=16, pady=(0, 16))

        def _riusa():
            self._riusa_query(entry["query"])
            dlg.destroy()

        ctk.CTkButton(
            bar, text="↺  Re-esegui", height=36,
            font=_font(11, "bold"), fg_color=ACCENT, hover_color="#00b386",
            text_color="#000000", corner_radius=10, command=_riusa,
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            bar, text="Chiudi", height=36,
            font=_font(11), fg_color=BTN_IDLE, hover_color=BTN_HOV,
            text_color=TEXT_SEC, border_color=BTN_BORDER, border_width=1,
            corner_radius=10, command=dlg.destroy,
        ).pack(side="right")

        dlg.bind("<Escape>", lambda _e: dlg.destroy())

    def _riusa_query(self, query: str):
        if self._busy or self._dataset is None:
            return
        self._txt_query.configure(state="normal")
        self._txt_query.delete(0, "end")
        self._txt_query.insert(0, query)
        self._txt_query.focus()

    # ── Export PDF ────────────────────────────────────────────────────────────

    def _esporta_pdf_sessione(self):
        if not self._report_items:
            self._msg_assistente(
                "Nessuna analisi riuscita da esportare.\n"
                "Fai almeno una domanda prima di generare il report.", tag="warn")
            return
        try:
            from fpdf import FPDF  # noqa: F401
        except ImportError:
            self._msg_assistente(
                "Libreria fpdf2 non installata.\n"
                "Esegui nel terminale:\n  pip install fpdf2 --break-system-packages",
                tag="err")
            return

        from tkinter import filedialog as _fd
        percorso = _fd.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf"), ("Tutti i file", "*.*")],
            title="Esporta Report PDF",
            initialfile=f"DataWhisperer_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        )
        if not percorso:
            return
        try:
            self._genera_pdf(percorso)
            self._msg_sistema(f"✓ Report esportato: {os.path.basename(percorso)}")
        except Exception as e:
            self._msg_assistente(f"Errore generazione PDF: {e}", tag="err")

    def _genera_pdf(self, percorso: str):
        import re as _re
        from fpdf import FPDF

        # ── helpers ──────────────────────────────────────────────────────────

        def _san(t: str) -> str:
            t = str(t)
            t = t.replace('—', '-').replace('–', '-')
            t = t.replace('‘', "'").replace('’', "'")
            t = t.replace('“', '"').replace('”', '"')
            t = t.replace('…', '...').replace('•', '-')
            t = t.replace('★', '*').replace('→', '->')
            t = t.replace('\xd7', 'x')
            return t.encode('latin-1', 'replace').decode('latin-1')

        def _no_md(t: str) -> str:
            t = _re.sub(r'\*\*(.+?)\*\*', r'\1', t)
            t = _re.sub(r'\*(.+?)\*',     r'\1', t)
            return t

        def _codice_breve(codice: str) -> str:
            righe, vuote = [], 0
            for r in codice.splitlines():
                if r.strip().startswith('#'):
                    continue
                if r.strip() == '':
                    vuote += 1
                    if vuote <= 1:
                        righe.append(r)
                else:
                    vuote = 0
                    righe.append(r)
            return '\n'.join(righe).strip()

        # ── palette (R, G, B) ────────────────────────────────────────────────
        C_TEAL   = (0,   200, 150)
        C_DARK   = (9,   9,   15)
        C_SLATE  = (78,  84,  110)
        C_IVORY  = (248, 250, 252)
        C_BORDER = (210, 220, 235)

        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=22)
        pdf.set_margins(18, 18, 18)
        pdf.alias_nb_pages()

        # ══════════════════════════════════════════════════════════════════════
        # COVER
        # ══════════════════════════════════════════════════════════════════════
        pdf.add_page()

        pdf.set_fill_color(*C_DARK)
        pdf.rect(0, 0, 210, 62, 'F')

        pdf.set_y(14)
        pdf.set_font('Helvetica', 'B', 30)
        pdf.set_text_color(*C_TEAL)
        pdf.cell(0, 14, 'Data-Whisperer', align='C', ln=True)

        pdf.set_font('Helvetica', '', 11)
        pdf.set_text_color(160, 170, 185)
        pdf.cell(0, 6, 'Report di Analisi Dati', align='C', ln=True)

        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(90, 100, 120)
        pdf.cell(0, 5, 'Air-Gapped  -  Zero rete  -  Dati elaborati localmente',
                 align='C', ln=True)

        # Card metadata
        pdf.ln(12)
        cy = pdf.get_y()
        pdf.set_fill_color(*C_IVORY)
        pdf.set_draw_color(*C_BORDER)
        pdf.set_line_width(0.3)
        pdf.rect(18, cy, 174, 34, 'FD')

        pdf.set_y(cy + 5)
        if self._dataset:
            pdf.set_font('Helvetica', 'B', 11)
            pdf.set_text_color(30, 40, 60)
            pdf.cell(0, 6, _san(self._dataset.nome_file), align='C', ln=True)
            pdf.set_font('Helvetica', '', 9)
            pdf.set_text_color(*C_SLATE)
            pdf.cell(0, 5,
                     _san(f'{self._dataset.n_righe} righe  |  {self._dataset.n_colonne} colonne'),
                     align='C', ln=True)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(*C_SLATE)
        pdf.cell(0, 5,
                 _san(f"Esportato il: {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
                      f"   |   Analisi nel report: {len(self._report_items)}"),
                 align='C', ln=True)

        # Sommario domande
        pdf.ln(10)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(*C_SLATE)
        pdf.cell(0, 5, 'SOMMARIO', ln=True)
        pdf.set_draw_color(*C_BORDER)
        pdf.set_line_width(0.2)
        pdf.line(18, pdf.get_y(), 192, pdf.get_y())
        pdf.ln(2)

        for i, item in enumerate(self._report_items, 1):
            if i % 2 == 0:
                pdf.set_fill_color(248, 250, 253)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(*C_TEAL)
            pdf.cell(10, 5.5, f'#{i}', fill=True)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(40, 50, 70)
            q = _san(item['query'])
            domanda_breve = q[:90] + ('...' if len(q) > 90 else '')
            pdf.cell(0, 5.5, domanda_breve, fill=True, ln=True)

        # ══════════════════════════════════════════════════════════════════════
        # PAGINE ANALISI
        # ══════════════════════════════════════════════════════════════════════
        for idx, item in enumerate(self._report_items, 1):
            pdf.add_page()

            # Header sezione
            pdf.set_fill_color(*C_DARK)
            pdf.set_text_color(*C_TEAL)
            pdf.set_font('Helvetica', 'B', 12)
            pdf.cell(0, 9, _san(f"  Analisi #{idx}   /   {item['ts']}"), fill=True, ln=True)
            pdf.ln(5)

            # Domanda
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_text_color(*C_TEAL)
            pdf.cell(0, 4, 'DOMANDA', ln=True)

            pdf.set_draw_color(*C_TEAL)
            pdf.set_line_width(0.4)
            y0 = pdf.get_y()
            pdf.set_fill_color(240, 252, 247)
            pdf.set_font('Helvetica', '', 11)
            pdf.set_text_color(20, 35, 30)
            pdf.multi_cell(0, 5.5, _san(item['query']), fill=True)
            pdf.rect(18, y0, 0.4, pdf.get_y() - y0, 'F')
            pdf.set_line_width(0.2)
            pdf.ln(6)

            # Risultato testuale
            _out = (item.get('output_testo') or '').strip()
            if _out and _out.lower() != 'none':
                pdf.set_font('Helvetica', 'B', 7)
                pdf.set_text_color(*C_SLATE)
                pdf.cell(0, 4, 'RISPOSTA', ln=True)
                pdf.set_fill_color(*C_IVORY)
                pdf.set_font('Helvetica', '', 10)
                pdf.set_text_color(30, 40, 55)
                pdf.multi_cell(0, 5, _san(_out), fill=True)
                pdf.ln(5)

            # Grafici centrati a 140 mm
            for png_bytes in item.get('grafici_png', []):
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                    tmp.write(png_bytes)
                    tmp_path = tmp.name
                try:
                    chart_w = 140.0
                    x_off = (210 - chart_w) / 2
                    pdf.image(tmp_path, x=x_off, w=chart_w)
                    pdf.ln(4)
                except Exception:
                    pass
                finally:
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass

            # Codice compatto senza commenti
            if item.get('codice'):
                codice_pulito = _codice_breve(_san(item['codice']))
                if codice_pulito:
                    pdf.set_font('Helvetica', 'B', 7)
                    pdf.set_text_color(*C_SLATE)
                    pdf.cell(0, 4, 'CODICE ESEGUITO', ln=True)
                    pdf.set_fill_color(243, 244, 250)
                    pdf.set_font('Courier', '', 7)
                    pdf.set_text_color(20, 90, 90)
                    pdf.multi_cell(0, 3.6, codice_pulito, fill=True)
                    pdf.ln(3)

            # Nota QA
            n_qa = item.get('tentativi_qa', 0)
            if n_qa > 0:
                pdf.set_font('Helvetica', 'I', 7)
                pdf.set_text_color(160, 130, 60)
                pdf.cell(0, 4,
                         _san(f"Auto-corretto in {n_qa} tentativo{'i' if n_qa > 1 else ''} dal QA."),
                         ln=True)

            # Numero pagina in fondo
            pdf.set_y(-14)
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(*C_SLATE)
            pdf.cell(0, 4, f'Pag. {pdf.page_no()}/{{nb}}  -  Data-Whisperer  -  Air-Gapped',
                     align='C')

        pdf.output(percorso)

    # ── Export XLSX ───────────────────────────────────────────────────────────

    def _esporta_xlsx_sessione(self):
        if not self._report_items:
            self._msg_assistente(
                "Nessuna analisi riuscita da esportare.", tag="warn")
            return
        from tkinter import filedialog as _fd
        percorso = _fd.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Tutti i file", "*.*")],
            title="Esporta XLSX",
            initialfile=f"DataWhisperer_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not percorso:
            return
        try:
            self._genera_xlsx(percorso)
            self._msg_sistema(f"✓ XLSX esportato: {os.path.basename(percorso)}")
        except Exception as e:
            self._msg_assistente(f"Errore generazione XLSX: {e}", tag="err")

    def _genera_xlsx(self, percorso: str):
        import io as _io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as _XlImage

        wb = Workbook()
        ws = wb.active
        ws.title = "Sommario"

        HDR_FILL  = PatternFill("solid", fgColor="09090F")
        HDR_FONT  = Font(name="Calibri", bold=True, color="00C896", size=11)
        BODY_FONT = Font(name="Calibri", color="1A2035", size=10)
        WRAP      = Alignment(wrap_text=True, vertical="top")
        CENTER    = Alignment(horizontal="center", vertical="center")

        headers = ["#", "Data/Ora", "Dataset", "Domanda", "Risposta", "Grafici"]
        ws.append(headers)
        for ci, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER

        for i, item in enumerate(self._report_items, 1):
            n_g = len(item.get("grafici_png", []))
            ws.append([
                i,
                item.get("ts", ""),
                self._dataset.nome_file if self._dataset else "",
                item.get("query", ""),
                (item.get("output_testo") or "")[:500],
                n_g,
            ])
            for ci in range(1, 7):
                cell = ws.cell(row=ws.max_row, column=ci)
                cell.font  = BODY_FONT
                cell.alignment = WRAP

        for ci, width in enumerate([4, 14, 22, 52, 62, 8], 1):
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # One sheet per chart
        g_idx = 0
        for item in self._report_items:
            for png_bytes in item.get("grafici_png", []):
                g_idx += 1
                ws_g = wb.create_sheet(title=f"Grafico {g_idx}")
                ws_g.cell(row=1, column=1).value = item.get("query", "")
                ws_g.cell(row=1, column=1).font  = Font(name="Calibri", bold=True, size=10)
                ws_g.column_dimensions["A"].width = 80

                img_stream = _io.BytesIO(png_bytes)
                xl_img     = _XlImage(img_stream)
                scale      = min(600 / max(xl_img.width, 1), 400 / max(xl_img.height, 1), 1.0)
                xl_img.width  = int(xl_img.width  * scale)
                xl_img.height = int(xl_img.height * scale)
                xl_img.anchor = "A3"
                ws_g.add_image(xl_img)

        wb.save(percorso)


